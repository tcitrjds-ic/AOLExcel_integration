"""CSV選考データマージツール

3つのD&D枠（2次選考 / 最終選考 / 内定出し）にCSVをドロップし、
各CSVから「日時」と「氏名」の2列を抜き出して、種別を付けた1つのCSVに集約する。
"""
from __future__ import annotations

import json
import os
import re
import subprocess
import sys
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

import pandas as pd

try:
    import openpyxl
except ImportError:
    print("openpyxl が未インストールです。`pip install openpyxl` を実行してください。")
    sys.exit(1)

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
except ImportError:
    print("tkinterdnd2 が未インストールです。`pip install tkinterdnd2` を実行してください。")
    sys.exit(1)


FULL_WIDTH_SPACE = "　"
TEMPLATE_SHEET_NAME = "送信データ"

APP_NAME = "CSV選考データマージ"
SETTINGS_FILENAME = "settings.json"

DEFAULT_EMPLOYEE_CSV = "社員一覧.csv"
DEFAULT_SHOZOKU_XLSX = "所属一覧.xlsx"
# 本部→To/Cc 宛先マスタ。運用上のファイル名は「本部一覧.xlsx」（旧称 AOL_ToとCc一覧.xlsx）。
DEFAULT_AOL_XLSX = "本部一覧.xlsx"
DEFAULT_TEMPLATE_XLSX = "個別送信テンプレート_AOL.xlsx"

# 所属一覧.xlsx（所属名称↔本部名称の対応表）の見出し名。
# 社員一覧 B列の所属名称（例: 浜松本部校）と AOL ToとCc一覧 の本部名称（例: 浜松本部）は
# 表記が一致しないため、この対応表で 所属名称 → 本部名称 へ変換して橋渡しする。
SHOZOKU_DEPT_HEADER = "所属名称"
SHOZOKU_HONBU_HEADER = "本部名称"


SECTIONS = [
    {
        "name": "2次選考",
        "date_col_name": "2次選考",
        "date_col_index": 2,   # C列
        "name_col_name": "グループ選考 【総合評価】（ユーザ）",
        "name_col_index": 7,   # H列
    },
    {
        "name": "最終選考",
        "date_col_name": "最終選考会",
        "date_col_index": 3,   # D列
        "name_col_name": "最終 【総合評価】（ユーザ）",
        "name_col_index": 9,   # J列
    },
    {
        "name": "内定出し",
        "date_col_name": "内定だし（＝面談）",
        "date_col_index": 4,   # E列
        "name_col_name": "内定通知 1.入社させたいか（ユーザ）",
        "name_col_index": 11,  # L列
    },
]

ENCODING_CANDIDATES = ["utf-8-sig", "cp932", "utf-8", "latin-1"]

# 元データのセル例: "合格 2026-03-04 13:00 WEB【1】" / "予約 2026-03-24 13:00 WEB【1】"
# 文字列中の "YYYY-M-D HH:MM"（区切りは - / 年月日いずれも許容）を抽出する。
DATETIME_PATTERN = re.compile(
    r"(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})日?\s*[T\s]\s*(\d{1,2})[:時](\d{2})"
)

# 氏名の姓・名区切り: 半角または全角スペース（連続は1区切りとみなす）
NAME_SPLIT_PATTERN = re.compile(r"[ 　]+")


def read_csv_auto(path: Path) -> pd.DataFrame:
    """エンコーディング候補を順に試してCSVを読み込む。"""
    last_err: Exception | None = None
    for enc in ENCODING_CANDIDATES:
        try:
            return pd.read_csv(path, encoding=enc, dtype=str, keep_default_na=False)
        except (UnicodeDecodeError, UnicodeError) as e:
            last_err = e
            continue
        except Exception as e:
            last_err = e
            continue
    raise RuntimeError(f"CSVの読み込みに失敗しました: {path.name} ({last_err})")


def format_datetime(value) -> str:
    """文字列から日時部分を抽出し `yyyy/M/d HH:mm` 形式に整形。

    元データのセルは "合格 2026-03-04 13:00 WEB【1】" のようにステータス語や場所名と
    一緒に格納されているため、まず正規表現で日時部分のみを抜き出す。
    日時が含まれない場合は空文字を返す。時刻の値（年月日・時分）は絶対に変更しない。
    """
    if value is None:
        return ""
    s = str(value).strip()
    if s == "" or s.lower() == "nan":
        return ""
    m = DATETIME_PATTERN.search(s)
    if m:
        year, month, day, hour, minute = m.groups()
        return f"{int(year)}/{int(month)}/{int(day)} {int(hour):02d}:{minute}"
    # フォールバック: 文字列全体を pandas でパース（純粋な日時文字列向け）
    parsed = pd.to_datetime(s, errors="coerce")
    if pd.isna(parsed):
        return ""
    return f"{parsed.year}/{parsed.month}/{parsed.day} {parsed.strftime('%H:%M')}"


def split_name(value) -> tuple[str, str]:
    """氏名を姓と名に分割する。

    - 区切り（半角/全角スペース）の最初の出現位置のみで分割
    - 区切りが見つからなければ全文字を姓に格納し、名は空（データ脱落防止）
    - 空文字や None は ("", "") を返す
    """
    if value is None:
        return ("", "")
    s = str(value).strip()
    if s == "" or s.lower() == "nan":
        return ("", "")
    parts = NAME_SPLIT_PATTERN.split(s, maxsplit=1)
    if len(parts) == 1:
        return (parts[0], "")
    return (parts[0].strip(), parts[1].strip())


def pick_column(df: pd.DataFrame, name: str, idx: int) -> pd.Series | None:
    """列名優先、見つからなければ位置（インデックス）でフォールバック。"""
    if name in df.columns:
        return df[name]
    if 0 <= idx < df.shape[1]:
        return df.iloc[:, idx]
    return None


def extract_section(files: list[Path], section: dict) -> tuple[pd.DataFrame, list[str]]:
    """指定セクションのCSV群から日時・氏名を抽出して長形式DataFrameを返す。

    戻り値: (DataFrame[日時, 氏名, 種別], 警告メッセージ一覧)
    """
    warnings: list[str] = []
    rows: list[pd.DataFrame] = []

    for path in files:
        try:
            df = read_csv_auto(path)
        except Exception as e:
            warnings.append(f"[{section['name']}] 読み込み失敗: {path.name} — {e}")
            continue

        if df.empty:
            warnings.append(f"[{section['name']}] 空のCSV: {path.name}")
            continue

        date_series = pick_column(df, section["date_col_name"], section["date_col_index"])
        name_series = pick_column(df, section["name_col_name"], section["name_col_index"])

        if date_series is None and name_series is None:
            warnings.append(
                f"[{section['name']}] 対象列が見つかりません（列名・位置とも不一致）: {path.name}"
            )
            continue

        if date_series is None:
            date_series = pd.Series([""] * len(df))
        if name_series is None:
            name_series = pd.Series([""] * len(df))

        # 氏名を姓・名に分割
        name_pairs = name_series.astype(str).map(split_name)
        sei_series = name_pairs.map(lambda t: t[0])
        mei_series = name_pairs.map(lambda t: t[1])

        sub = pd.DataFrame({
            "日時": date_series.astype(str).map(format_datetime),
            "姓": sei_series,
            "名": mei_series,
        })
        # 姓・名がともに空の行は除外（=元の氏名が空白）
        sub = sub[~((sub["姓"] == "") & (sub["名"] == ""))]
        sub["種別"] = section["name"]
        rows.append(sub)

    if not rows:
        return pd.DataFrame(columns=["日時", "姓", "名", "種別"]), warnings
    return pd.concat(rows, ignore_index=True), warnings


def load_employee_directory(path: Path) -> dict[str, str]:
    """社員一覧.csv → {社員名: 所属} の辞書。

    社員一覧は列0=社員名, 列1=所属 の固定レイアウト。社員名は「姓　名」(全角空白) で
    格納されているのが基本だが、表記揺れ対策として半角空白版もキーとして登録する。
    """
    df = read_csv_auto(path)
    if df.shape[1] < 2:
        raise RuntimeError(f"社員一覧の列数が不足しています: {path.name}")

    result: dict[str, str] = {}
    name_col = df.iloc[:, 0]
    dept_col = df.iloc[:, 1]
    for name, dept in zip(name_col, dept_col):
        n = str(name).strip()
        if not n or n.lower() == "nan":
            continue
        d = str(dept).strip()
        if d.lower() == "nan":
            d = ""
        result.setdefault(n, d)
        # 区切り表記揺れに対する保険
        alt_full = n.replace(" ", FULL_WIDTH_SPACE)
        if alt_full != n:
            result.setdefault(alt_full, d)
        alt_half = n.replace(FULL_WIDTH_SPACE, " ")
        if alt_half != n:
            result.setdefault(alt_half, d)
    return result


def load_shozoku_map(path: Path) -> dict[str, str]:
    """所属一覧.xlsx → {所属名称: 本部名称} の変換テーブル。

    社員一覧 B列の所属名称は AOL ToとCc一覧 の本部名称と表記が一致しないため、
    この対応表で 所属名称 → 本部名称 へ橋渡しする（D→Bの逆引きではなく所属→本部の前方変換）。

    対象シートは見出しに「所属名称」「本部名称」を両方持つシートを自動選択する
    （同ファイルに To/Cc一覧のシートが同梱されている場合があるため）。見つからなければ
    アクティブシートを使い、列は見出し名で特定、無ければ B列=所属名称 / D列=本部名称 で
    位置フォールバックする。
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    def header_of(ws) -> list[str]:
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            return ["" if c is None else str(c).strip() for c in row]
        return []

    target_ws = None
    for ws in wb.worksheets:
        hdr = header_of(ws)
        if SHOZOKU_DEPT_HEADER in hdr and SHOZOKU_HONBU_HEADER in hdr:
            target_ws = ws
            break
    if target_ws is None:
        target_ws = wb.active

    hdr = header_of(target_ws)
    dept_idx = hdr.index(SHOZOKU_DEPT_HEADER) if SHOZOKU_DEPT_HEADER in hdr else 1
    honbu_idx = hdr.index(SHOZOKU_HONBU_HEADER) if SHOZOKU_HONBU_HEADER in hdr else 3

    result: dict[str, str] = {}
    for row in target_ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        dept = row[dept_idx] if len(row) > dept_idx else None
        honbu = row[honbu_idx] if len(row) > honbu_idx else None
        d = str(dept).strip() if dept is not None else ""
        h = str(honbu).strip() if honbu is not None else ""
        if not d or d.lower() == "nan":
            continue
        if not h or h.lower() == "nan":
            continue
        result.setdefault(d, h)
    return result


def load_aol_to_cc(path: Path) -> dict[str, tuple[list[str], list[str]]]:
    """AOL_ToとCc一覧.xlsx → {本部名: (To氏名list, Cc氏名list)}。

    A=本部コード, B=本部名 は本部の先頭行だけ埋まる縦持ち。同本部に属する追加の
    To/Cc行はA,Bが空欄で並ぶため、走査中の current_branch に紐付ける。
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    result: dict[str, tuple[list[str], list[str]]] = {}
    current_branch: str | None = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        name = row[1] if len(row) > 1 else None
        to_name = row[2] if len(row) > 2 else None
        cc_name = row[3] if len(row) > 3 else None

        if name is not None:
            n = str(name).strip()
            if n:
                current_branch = n
                result.setdefault(current_branch, ([], []))

        if current_branch is None:
            continue

        if to_name is not None:
            tn = str(to_name).strip()
            if tn:
                result[current_branch][0].append(tn)
        if cc_name is not None:
            cn = str(cc_name).strip()
            if cn:
                result[current_branch][1].append(cn)
    return result


# 連絡先オーバーライドの複数名区切り（「古田直樹、大久保佑亮」等）
CONTACT_NAME_SPLIT_PATTERN = re.compile(r"[、,，/／・\n\r]+")


def _norm_name(value) -> str:
    """氏名・本部名比較用の正規化: 前後空白除去 + 全角/半角スペース除去。"""
    if value is None:
        return ""
    return str(value).strip().replace(FULL_WIDTH_SPACE, "").replace(" ", "")


def _split_people(cell) -> list[str]:
    """連絡先セル「古田直樹、大久保佑亮」→ ['古田直樹', '大久保佑亮']。空要素は除去。"""
    if cell is None:
        return []
    s = str(cell).strip()
    if not s or s.lower() == "nan":
        return []
    return [p.strip() for p in CONTACT_NAME_SPLIT_PATTERN.split(s) if p.strip()]


def load_contact_override(
    path: Path,
) -> tuple[dict[str, tuple[list[str], list[str]]], dict[str, tuple[list[str], list[str]]]]:
    """「正確な連絡先」xlsx を読み、(氏名キー, 本部/所属キー) の宛先オーバーライド2種を返す。

    想定レイアウト（ヘッダ名で列を判定。順不同・余剰列可）:
        本部 | 所属 | 氏名(任意) | TO | CC
    - TO/CC は「古田直樹、大久保佑亮」のように 、 区切りで複数名。
    - 氏名列がある版は対象者個人に宛先を固定でき、社員一覧の所属ズレ
      （例: 社員一覧上は磐田だが実際は掛川で選考に立つ）も修正できる（氏名優先）。
    - 氏名列が無い版は本部/所属キーのみ生成（本部一覧の宛先差し替え用）。

    戻り値:
        by_name: {正規化氏名: (to_list, cc_list)}
        by_dept: {正規化(本部名称 or 所属名称): (to_list, cc_list)}  ※本部・所属の両方をキー登録
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}, {}

    header = [_norm_name(h).upper() for h in rows[0]]

    def find_col(cands: set[str]) -> int | None:
        for i, h in enumerate(header):
            if h in cands:
                return i
        return None

    ci_honbu = find_col({"本部", "本部名", "本部名称"})
    ci_shoz = find_col({"所属", "所属名", "所属名称"})
    ci_name = find_col({"氏名", "対象者", "名前"})
    ci_to = find_col({"TO"})
    ci_cc = find_col({"CC"})
    if ci_to is None and ci_cc is None:
        raise RuntimeError(f"連絡先ファイルに TO/CC 列が見つかりません: {path.name}")

    def cell(row, idx):
        return row[idx] if (idx is not None and idx < len(row)) else None

    by_name: dict[str, tuple[list[str], list[str]]] = {}
    by_dept: dict[str, tuple[list[str], list[str]]] = {}
    for row in rows[1:]:
        if row is None:
            continue
        to_list = _split_people(cell(row, ci_to))
        cc_list = _split_people(cell(row, ci_cc))
        if not to_list and not cc_list:
            continue
        nm = _norm_name(cell(row, ci_name)) if ci_name is not None else ""
        if nm:
            by_name.setdefault(nm, (to_list, cc_list))
        for key_raw in (cell(row, ci_honbu), cell(row, ci_shoz)):
            k = _norm_name(key_raw)
            if k:
                by_dept.setdefault(k, (to_list, cc_list))
    return by_name, by_dept


def _build_name_speller(
    emp_dir: dict[str, str],
    aol_map: dict[str, tuple[list[str], list[str]]],
) -> dict[str, str]:
    """空白無し氏名 → 社員一覧/本部一覧にある空白付き正式表記 の対応表。

    連絡先ファイルの氏名は「田丸敏之」のように姓名間の空白が無いことが多い。送信CSVの
    姓/名分割・表記統一のため、既知の空白付き表記（例「田丸　敏之」）へ可能なら戻す。
    """
    spell: dict[str, str] = {}

    def add(name) -> None:
        n = ("" if name is None else str(name)).strip()
        key = _norm_name(n)
        if not key:
            return
        has_sp = (FULL_WIDTH_SPACE in n) or (" " in n)
        if key not in spell:
            spell[key] = n
        elif has_sp and not ((FULL_WIDTH_SPACE in spell[key]) or (" " in spell[key])):
            spell[key] = n  # 空白付き表記を優先

    for k in emp_dir.keys():
        add(k)
    for to_list, cc_list in aol_map.values():
        for n in to_list:
            add(n)
        for n in cc_list:
            add(n)
    return spell


def load_template_defaults(path: Path) -> dict[str, str]:
    """個別送信テンプレ.xlsx の2行目から件名/本文/添付/BCCの既定値を取得。

    3行目は各列の入力規約説明のため読み飛ばす。
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    if TEMPLATE_SHEET_NAME in wb.sheetnames:
        ws = wb[TEMPLATE_SHEET_NAME]
    else:
        ws = wb.active

    def cell(coord: str) -> str:
        v = ws[coord].value
        return "" if v is None else str(v)

    return {
        "件名": cell("C2"),
        "本文": cell("D2"),
        "添付ファイル名": cell("E2"),
        "BCC氏名": cell("G2"),
    }


def compose_cc_names(names_in_order: list[str], exclude: set[str] | None = None) -> str:
    """CC氏名文字列を構築する。

    `names_in_order` を順に走査し、空白を除去・出現順保持で重複除去・`exclude` 一致を弾いて、
    半角カンマで結合する。
    """
    seen: set[str] = set()
    if exclude:
        for ex in exclude:
            s = ex.strip()
            if s:
                seen.add(s)
    ordered: list[str] = []
    for name in names_in_order:
        n = name.strip()
        if not n or n in seen:
            continue
        seen.add(n)
        ordered.append(n)
    return ",".join(ordered)


SEND_CSV_COLUMNS = [
    "姓",
    "名",
    "件名",
    "本文",
    "添付ファイル名",
    "CC氏名",
    "BCC氏名",
    "選考日時",
    "種別",
]

# 選考日時の時刻変換ルール（運用上、選考時刻は HH:00 固定で来る前提）。
# key  : マージCSV由来の時刻部分（半角コロン）
# value: テンプレ「選考日時」列に書き出す表示文字列（全角コロン・全角チルダ）
TIME_SLOT_MAP: dict[str, dict[str, str]] = {
    "2次選考": {
        "13:00": "13：45～15：00",
    },
    "最終選考": {
        "14:00": "14：00～15：00",
        "15:00": "15：00～16：00",
    },
    "内定出し": {
        "14:00": "14：00～14：50",
        "15:00": "15：00～15：50",
    },
}

# 最終選考: 同一(担当者・日付)で 14:00 と 15:00 の両方が実施されている場合の統合表示
FINAL_COMBINED_BOTH_SLOTS = "14：00～16：00"

# 「人事で登録する稲毛執行役・吉留先生・中萬執行役の分」に該当する所属。
# 該当行はAOL送信CSVに出力しない（未マッチログには記録）。
EXCLUDED_DEPTS: set[str] = {
    "さなる東京総本社",
    "さなる東京総本社（テストあり）",
    "さなる名古屋本社",
    # 「東京本社」相当（社員一覧上は機能別に分かれているが、人事登録分として扱う）
    "東京広報室",
    "東京教材研究室",
    # 「名古屋本社」相当（業務ブロックは除く本社機能の所属）
    "名古屋広報室",
    "名古屋撮影室",
    "名古屋教材研究室",
}

# 人事で別途登録する3名（社員一覧で姓と1:1対応していることを確認済み）。
# 稲毛 重典 (稲毛執行役) / 吉留 博巳 (吉留先生) / 中萬 隆信 (中萬執行役)
# 該当者は未マッチにも出さず静かに除外する（出力もログも対象外）。
EXCLUDED_LAST_NAMES: set[str] = {
    "稲毛",
    "吉留",
    "中萬",
}

# 人事が別途登録するため除外する所属（完全一致）。未マッチにも出さず静かに除外する。
# ※「スクール21人事室」「中萬学院人事室」等は対象外（"人事室" 完全一致のみ）。
EXCLUDED_DEPTS_SILENT: set[str] = {
    "人事室",
}


def _split_date_time(dt_str: str) -> tuple[str, str]:
    """'2026/3/4 13:00' → ('2026/3/4', '13:00')。空白で1回だけ分割。"""
    s = (dt_str or "").strip()
    parts = s.split(" ", 1)
    if len(parts) == 2:
        return parts[0].strip(), parts[1].strip()
    return s, ""


def _make_send_row(entry: dict, selection_dt: str, defaults: dict[str, str]) -> dict:
    """1件の解決済みエントリから送信CSV1行を構築する。

    [TODO/暫定] 現行送信システムは To 1名のみ。AOL.To 先頭1名をA/B(姓・名)、残りはCC氏名へ。
    将来複数To対応時に 1メール=1人の行複製版（git履歴）へ戻す。
    """
    to_list = entry["to_list"]
    cc_list = entry["cc_list"]
    primary_to = to_list[0]
    rest_to = to_list[1:]
    to_sei, to_mei = split_name(primary_to)
    cc_names = compose_cc_names(
        [entry["full"]] + rest_to + cc_list,
        exclude={primary_to},
    )
    return {
        "姓": to_sei,
        "名": to_mei,
        "件名": defaults.get("件名", ""),
        "本文": defaults.get("本文", ""),
        "添付ファイル名": defaults.get("添付ファイル名", ""),
        "CC氏名": cc_names,
        "BCC氏名": defaults.get("BCC氏名", ""),
        "選考日時": selection_dt,
        "種別": entry["shubetsu"],
    }


def build_send_rows(
    merged_df: pd.DataFrame,
    emp_dir: dict[str, str],
    shozoku_map: dict[str, str],
    aol_map: dict[str, tuple[list[str], list[str]]],
    defaults: dict[str, str],
    contact_override: tuple[dict, dict] | None = None,
) -> tuple[pd.DataFrame, list[dict]]:
    """マージ済みDataFrameをAOL送信用DataFrameに変換し、未マッチ詳細も返す。

    宛先(To/Cc)の決定順（contact_override 指定時は①②が優先）:
        ① 連絡先・氏名一致: 対象者氏名が連絡先ファイルに在れば、社員一覧/本部一覧を
           介さずそのTo/Ccを採用（社員一覧の所属ズレ・各種除外より優先）。
        ② 連絡先・本部/所属一致: 解決した所属名称または本部名称が連絡先ファイルに
           在ればそのTo/Ccで本部一覧を差し替え。
        ③ 既定: 社員一覧→所属名称→（所属一覧で本部名称へ変換）→本部一覧で照合。

    その他の処理順:
        1. 行ごとに 社員一覧→所属名称、所属除外、所属一覧で所属名称→本部名称へ変換、
           AOL本部マスタ照合、To人物存在チェック
        2. 最終選考のみ、同一(担当者・日付)に 14:00 と 15:00 両方ある場合は1行に統合
           （選考日時 = "14：00～16：00"）
        3. それ以外は TIME_SLOT_MAP で時刻を窓表記に変換して1行出力
        4. 変換ルール未定義の時刻はスキップしてログ集約
    """
    by_name, by_dept = contact_override if contact_override else ({}, {})
    speller = _build_name_speller(emp_dir, aol_map) if (by_name or by_dept) else {}

    def spell_all(names: list[str]) -> list[str]:
        return [speller.get(_norm_name(n), n) for n in names]

    unmatched: list[dict] = []
    resolved: list[dict] = []

    for _, row in merged_df.iterrows():
        sei = str(row.get("姓", "")).strip()
        mei = str(row.get("名", "")).strip()
        shubetsu = str(row.get("種別", "")).strip()
        full_fwsp = f"{sei}{FULL_WIDTH_SPACE}{mei}" if mei else sei
        full_hwsp = f"{sei} {mei}" if mei else sei
        dt = str(row.get("日時", ""))

        # ① 連絡先・氏名一致（最優先）: 社員一覧/本部一覧・各種除外を介さず宛先を固定
        ov = by_name.get(_norm_name(full_fwsp)) if by_name else None
        if ov is not None:
            to_list = spell_all(ov[0])
            cc_list = spell_all(ov[1])
            if not to_list:
                unmatched.append({
                    "日時": dt, "氏名": full_fwsp, "種別": shubetsu,
                    "理由": "連絡先(氏名一致)にTo人物が居ない",
                })
                continue
            date_part, time_part = _split_date_time(dt)
            resolved.append({
                "sei": sei, "mei": mei, "full": full_fwsp,
                "dept": "(連絡先・氏名指定)", "honbu": "(連絡先・氏名指定)",
                "date": date_part, "time": time_part, "dt": dt,
                "shubetsu": shubetsu,
                "to_list": to_list, "cc_list": cc_list,
            })
            continue

        if sei in EXCLUDED_LAST_NAMES:
            # 人事で別途登録する分。未マッチにも出さず静かに除外する。
            continue

        dept = emp_dir.get(full_fwsp)
        if dept is None:
            dept = emp_dir.get(full_hwsp)
        if not dept:
            unmatched.append({"日時": dt, "氏名": full_fwsp, "種別": shubetsu, "理由": "社員一覧に該当なし"})
            continue

        if dept in EXCLUDED_DEPTS_SILENT:
            # 人事室など人事登録分。未マッチにも出さず静かに除外する。
            continue

        if dept in EXCLUDED_DEPTS:
            unmatched.append({
                "日時": dt, "氏名": full_fwsp, "種別": shubetsu,
                "理由": f"除外: 人事登録分（所属={dept}）",
            })
            continue

        # 所属一覧で 所属名称 → 本部名称 へ橋渡し（対応表に無ければ所属名をそのままキーに試す）
        honbu = shozoku_map.get(dept, dept)

        # ② 連絡先・本部/所属一致 → 本部一覧より優先で差し替え。無ければ ③ 既定の本部一覧。
        ov2 = None
        if by_dept:
            ov2 = by_dept.get(_norm_name(dept))
            if ov2 is None:
                ov2 = by_dept.get(_norm_name(honbu))
        if ov2 is not None:
            to_list, cc_list = spell_all(ov2[0]), spell_all(ov2[1])
        else:
            to_cc = aol_map.get(honbu)
            if to_cc is None:
                if dept in shozoku_map:
                    reason = f"AOL ToとCc一覧に本部「{honbu}」なし（所属「{dept}」より変換）"
                else:
                    reason = f"所属一覧に「{dept}」なし／AOL本部マスタにも該当なし"
                unmatched.append({
                    "日時": dt, "氏名": full_fwsp, "種別": shubetsu,
                    "理由": reason,
                })
                continue
            to_list, cc_list = to_cc
        if not to_list:
            unmatched.append({
                "日時": dt, "氏名": full_fwsp, "種別": shubetsu,
                "理由": f"「{honbu}」にTo人物が居ない",
            })
            continue

        date_part, time_part = _split_date_time(dt)
        resolved.append({
            "sei": sei, "mei": mei, "full": full_fwsp,
            "dept": dept, "honbu": honbu,
            "date": date_part, "time": time_part, "dt": dt,
            "shubetsu": shubetsu,
            "to_list": to_list, "cc_list": cc_list,
        })

    # 最終選考の14:00/15:00統合判定
    final_groups: dict[tuple, list[int]] = {}
    for i, e in enumerate(resolved):
        if e["shubetsu"] == "最終選考":
            final_groups.setdefault((e["full"], e["date"]), []).append(i)

    combined_into: set[int] = set()
    combined_primary: set[int] = set()
    for indices in final_groups.values():
        has_14 = any(resolved[i]["time"] == "14:00" for i in indices)
        has_15 = any(resolved[i]["time"] == "15:00" for i in indices)
        if has_14 and has_15:
            primary = next(i for i in indices if resolved[i]["time"] == "14:00")
            combined_primary.add(primary)
            for i in indices:
                if i != primary and resolved[i]["time"] in ("14:00", "15:00"):
                    combined_into.add(i)

    out_rows: list[dict] = []
    for i, e in enumerate(resolved):
        if i in combined_into:
            continue
        if i in combined_primary:
            new_dt = f"{e['date']} {FINAL_COMBINED_BOTH_SLOTS}"
            out_rows.append(_make_send_row(e, new_dt, defaults))
            continue

        mapping = TIME_SLOT_MAP.get(e["shubetsu"], {})
        converted = mapping.get(e["time"])
        if converted is None:
            unmatched.append({
                "日時": e["dt"], "氏名": e["full"], "種別": e["shubetsu"],
                "理由": f"時刻 {e['time']!r} の変換ルール無し（{e['shubetsu']}）",
            })
            continue
        new_dt = f"{e['date']} {converted}"
        out_rows.append(_make_send_row(e, new_dt, defaults))

    return pd.DataFrame(out_rows, columns=SEND_CSV_COLUMNS), unmatched


SKIPPED_LOG_COLUMNS = ["日時", "氏名", "種別", "理由"]


def write_skipped_log(unmatched: list[dict], out_path: Path) -> None:
    """未マッチ行（登録できなかった行）の一覧を CSV ファイルに書き出す。"""
    df = pd.DataFrame(unmatched, columns=SKIPPED_LOG_COLUMNS)
    df.to_csv(out_path, index=False, encoding="utf-8-sig")


def get_downloads_dir() -> Path:
    """Windowsのダウンロードフォルダのパスを返す（無ければホーム直下にフォールバック）。"""
    home = Path.home()
    downloads = home / "Downloads"
    downloads.mkdir(parents=True, exist_ok=True)
    return downloads


def resource_path(name: str) -> Path:
    """exe に同梱した参照ファイルの実パスを返す。

    PyInstaller でビルドした exe では同梱物が一時展開先 sys._MEIPASS に置かれる。
    開発実行時（未凍結）はソース隣の templates/ を参照する。これにより配布した単一 exe に
    参照ファイルを内蔵でき、絶対パスを直書きせずに「最初から読み込まれた状態」を実現する。
    """
    base = getattr(sys, "_MEIPASS", None)
    if base:
        return Path(base) / name
    return Path(__file__).resolve().parent / "templates" / name


def get_config_path() -> Path:
    """参照ファイルの選択を記憶する設定ファイルの保存先を返す。

    Windows の %APPDATA%\\CSV選考データマージ\\settings.json。APPDATA が無い環境では
    ホーム直下の .CSV選考データマージ にフォールバックする。保存先は実行時にユーザー環境から
    解決し、コードに絶対パスを直書きしない（ユーザー／PCが変わっても追従できる）。
    """
    base = os.environ.get("APPDATA")
    config_dir = Path(base) / APP_NAME if base else Path.home() / f".{APP_NAME}"
    config_dir.mkdir(parents=True, exist_ok=True)
    return config_dir / SETTINGS_FILENAME


def load_settings() -> dict:
    """設定ファイルを読み、{設定キー: 値} のdictを返す（無ければ空dict）。"""
    path = get_config_path()
    if not path.exists():
        return {}
    try:
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def save_settings(data: dict) -> None:
    """設定ファイルへ {設定キー: 値} を書き出す（失敗しても処理は止めない）。"""
    try:
        path = get_config_path()
        with path.open("w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def open_in_explorer(path: Path) -> None:
    """エクスプローラーで指定ファイルを選択状態で開く（Windows専用）。"""
    try:
        subprocess.Popen(["explorer", "/select,", str(path)])
    except Exception:
        # フォールバック: フォルダだけでも開く
        try:
            os.startfile(path.parent)
        except Exception:
            pass


class DropZone(ttk.LabelFrame):
    """D&D対応のCSV受領フレーム。"""

    def __init__(self, parent: tk.Widget, title: str):
        super().__init__(parent, text=f"＜{title}＞", padding=8)
        self.title_text = title
        self.files: list[Path] = []

        hint = ttk.Label(
            self,
            text="ここにCSVをドラッグ&ドロップ\n（複数可）",
            anchor="center",
            justify="center",
            foreground="#666",
        )
        hint.pack(fill="x", pady=(0, 4))

        self.listbox = tk.Listbox(self, height=10, activestyle="none")
        self.listbox.pack(fill="both", expand=True)

        btn_row = ttk.Frame(self)
        btn_row.pack(fill="x", pady=(6, 0))
        ttk.Button(btn_row, text="選択を削除", command=self._remove_selected).pack(side="left")
        ttk.Button(btn_row, text="全クリア", command=self._clear_all).pack(side="right")

        # D&D登録（フレーム本体とListbox両方）
        for widget in (self, hint, self.listbox):
            widget.drop_target_register(DND_FILES)
            widget.dnd_bind("<<Drop>>", self._on_drop)

    def _on_drop(self, event) -> str:
        # event.data はスペース区切り。空白を含むパスは{}で囲まれる。
        raw = event.data
        paths = self._parse_dnd_paths(raw)
        added = 0
        skipped: list[str] = []
        for p in paths:
            path = Path(p)
            if not path.exists() or path.suffix.lower() != ".csv":
                skipped.append(path.name)
                continue
            if path in self.files:
                continue
            self.files.append(path)
            self.listbox.insert("end", path.name)
            added += 1
        if skipped:
            messagebox.showinfo(
                "スキップ",
                f"CSV以外のファイルはスキップしました:\n" + "\n".join(skipped),
            )
        return "break"

    @staticmethod
    def _parse_dnd_paths(data: str) -> list[str]:
        """tkinterdnd2の<<Drop>>イベントdata文字列をパス一覧に変換。"""
        result: list[str] = []
        i = 0
        n = len(data)
        while i < n:
            if data[i] == " ":
                i += 1
                continue
            if data[i] == "{":
                j = data.find("}", i + 1)
                if j == -1:
                    result.append(data[i + 1:])
                    break
                result.append(data[i + 1:j])
                i = j + 1
            else:
                j = data.find(" ", i)
                if j == -1:
                    result.append(data[i:])
                    break
                result.append(data[i:j])
                i = j + 1
        return result

    def _remove_selected(self) -> None:
        for idx in reversed(self.listbox.curselection()):
            del self.files[idx]
            self.listbox.delete(idx)

    def _clear_all(self) -> None:
        self.files.clear()
        self.listbox.delete(0, "end")


class FilePickerRow(ttk.Frame):
    """単一ファイル選択行: ラベル + パス表示 + 参照ボタン + D&D対応。"""

    def __init__(
        self,
        parent: tk.Widget,
        label: str,
        extensions: tuple[str, ...],
        default_path: Path | None = None,
        clearable: bool = False,
    ):
        super().__init__(parent)
        self.label_text = label
        self.extensions = tuple(e.lower() for e in extensions)
        self.path: Path | None = None
        # パス変更時に呼ぶコールバック（設定の保存用）。構築後に App が代入する。
        # 構築時の default_path 適用では None のままなので保存は走らない。
        self.on_change = None

        ttk.Label(self, text=label, width=26, anchor="w").pack(side="left")
        self.var = tk.StringVar(value="(未選択)")
        self.entry = ttk.Entry(self, textvariable=self.var, state="readonly")
        self.entry.pack(side="left", fill="x", expand=True, padx=(0, 6))
        ttk.Button(self, text="参照...", command=self._browse).pack(side="left")
        if clearable:
            # 任意ファイル用: 選択を解除して未指定状態に戻せるようにする
            ttk.Button(self, text="クリア", command=self._clear).pack(side="left", padx=(4, 0))

        for widget in (self, self.entry):
            widget.drop_target_register(DND_FILES)
            widget.dnd_bind("<<Drop>>", self._on_drop)

        if default_path is not None and default_path.exists():
            self._set_path(default_path)

    def _browse(self) -> None:
        filetypes = [(f"{e.lstrip('.').upper()} files", f"*{e}") for e in self.extensions]
        filetypes.append(("All files", "*.*"))
        p = filedialog.askopenfilename(title=f"{self.label_text} を選択", filetypes=filetypes)
        if p:
            self._set_path(Path(p))

    def _on_drop(self, event) -> str:
        paths = DropZone._parse_dnd_paths(event.data)
        for p in paths:
            path = Path(p)
            if path.exists() and path.suffix.lower() in self.extensions:
                self._set_path(path)
                return "break"
        messagebox.showinfo(
            "スキップ",
            f"{self.label_text}: 対象拡張子 ({', '.join(self.extensions)}) のファイルが含まれていません。",
        )
        return "break"

    def _set_path(self, p: Path) -> None:
        self.path = p
        self.var.set(p.name)
        if self.on_change is not None:
            self.on_change()

    def _clear(self) -> None:
        self.path = None
        self.var.set("(未選択)")
        if self.on_change is not None:
            self.on_change()


class App:
    def __init__(self, root: TkinterDnD.Tk):
        self.root = root
        root.title("CSV選考データマージ")
        root.minsize(720, 400)

        container = ttk.Frame(root, padding=10)
        container.pack(fill="both", expand=True)

        zones_frame = ttk.Frame(container)
        zones_frame.pack(fill="both", expand=True)

        self.zones: list[DropZone] = []
        for i, sec in enumerate(SECTIONS):
            zone = DropZone(zones_frame, sec["name"])
            zone.grid(row=0, column=i, sticky="nsew", padx=4)
            zones_frame.columnconfigure(i, weight=1)
            self.zones.append(zone)
        zones_frame.rowconfigure(0, weight=1)

        # 前回選択したパスを設定ファイルから復元。優先順位は 保存済み → exe同梱 → Downloads配下。
        self.settings = load_settings()
        ref_frame = ttk.LabelFrame(
            container,
            text="リファレンスファイル（AOL送信CSV出力用 / exe同梱を初期値・前回の選択を記憶・GUIで差し替え可）",
            padding=8,
        )
        ref_frame.pack(fill="x", pady=(10, 0))
        self.emp_picker = FilePickerRow(
            ref_frame, "社員一覧 (CSV)", (".csv",),
            default_path=self._initial_ref_path("emp", DEFAULT_EMPLOYEE_CSV),
        )
        self.emp_picker.pack(fill="x", pady=2)
        self.shozoku_picker = FilePickerRow(
            ref_frame, "所属一覧 (XLSX)", (".xlsx",),
            default_path=self._initial_ref_path("shozoku", DEFAULT_SHOZOKU_XLSX),
        )
        self.shozoku_picker.pack(fill="x", pady=2)
        self.aol_picker = FilePickerRow(
            ref_frame, "本部一覧 (XLSX)", (".xlsx",),
            default_path=self._initial_ref_path("aol", DEFAULT_AOL_XLSX),
        )
        self.aol_picker.pack(fill="x", pady=2)
        self.tpl_picker = FilePickerRow(
            ref_frame, "個別送信テンプレート (XLSX)", (".xlsx",),
            default_path=self._initial_ref_path("tpl", DEFAULT_TEMPLATE_XLSX),
        )
        self.tpl_picker.pack(fill="x", pady=2)
        # 連絡先オーバーライド（任意）: 実施月ごとの「正確な連絡先」で宛先を上書きする。
        # 月ごとに中身が変わるファイルのため、exe同梱も設定保存も意図的にしない
        # （古い月の連絡先が翌月の出力へ無言で適用される事故を防ぐ）。クリアで解除可。
        self.contact_picker = FilePickerRow(
            ref_frame, "連絡先オーバーライド (XLSX/任意)", (".xlsx",),
            default_path=None,
            clearable=True,
        )
        self.contact_picker.pack(fill="x", pady=2)

        # 設定キー → ピッカー。パス変更時に全ピッカーの現在値を設定ファイルへ保存する。
        # contact_picker は上記の理由で意図的に含めない（保存・復元の対象外）。
        self._ref_pickers: dict[str, FilePickerRow] = {
            "emp": self.emp_picker,
            "shozoku": self.shozoku_picker,
            "aol": self.aol_picker,
            "tpl": self.tpl_picker,
        }
        for picker in self._ref_pickers.values():
            picker.on_change = self._persist_ref_paths

        btn_row = ttk.Frame(container)
        btn_row.pack(fill="x", pady=(10, 0))
        ttk.Button(btn_row, text="マージしてCSV出力", command=self.on_merge_click).pack(side="left", padx=4)
        ttk.Button(btn_row, text="AOL送信CSV出力", command=self.on_aol_send_click).pack(side="left", padx=4)

        self.status = ttk.Label(container, text="準備完了", foreground="#444")
        self.status.pack(fill="x", pady=(8, 0))

    def _initial_ref_path(self, key: str, filename: str) -> Path:
        """参照ファイルの初期パス。優先順位は 保存済み(実在) → exe同梱 → Downloads配下。

        絶対パスは直書きせず、いずれも実行時に解決する。配布された単一exeでは同梱コピーが
        初期値として読み込まれ、利用者はGUIで差し替え可能（差し替えは設定ファイルに保存）。
        """
        saved = self.settings.get(key)
        if saved:
            p = Path(saved)
            if p.exists():
                return p
        bundled = resource_path(filename)
        if bundled.exists():
            return bundled
        return get_downloads_dir() / filename

    def _persist_ref_paths(self) -> None:
        """全リファレンスピッカーの現在のパスを設定ファイルへ保存する。"""
        for key, picker in self._ref_pickers.items():
            if picker.path is not None:
                self.settings[key] = str(picker.path)
        save_settings(self.settings)

    def _merge_dropped_csvs(self) -> tuple[pd.DataFrame | None, list[str]]:
        """既存D&D 3枠の内容をマージし、(DataFrame or None, 警告list) を返す。"""
        all_files = [zone.files for zone in self.zones]
        if all(len(f) == 0 for f in all_files):
            messagebox.showwarning("ファイルなし", "少なくとも1つの枠にCSVをドロップしてください。")
            return None, []

        all_warnings: list[str] = []
        frames: list[pd.DataFrame] = []
        for zone, section in zip(self.zones, SECTIONS):
            if not zone.files:
                continue
            df, warnings = extract_section(zone.files, section)
            all_warnings.extend(warnings)
            if not df.empty:
                frames.append(df)

        if not frames:
            messagebox.showerror(
                "抽出失敗",
                "対象列が抽出できませんでした。\n\n" + "\n".join(all_warnings),
            )
            return None, all_warnings

        return pd.concat(frames, ignore_index=True), all_warnings

    def on_merge_click(self) -> None:
        merged, all_warnings = self._merge_dropped_csvs()
        if merged is None:
            return

        out_dir = get_downloads_dir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = out_dir / f"merged_{timestamp}.csv"

        try:
            merged.to_csv(out_path, index=False, encoding="utf-8-sig")
        except Exception as e:
            messagebox.showerror("書き出し失敗", f"CSVの書き出しに失敗しました:\n{e}")
            return

        self.status.config(text=f"出力: {out_path}")

        warn_text = "\n\n警告:\n" + "\n".join(all_warnings) if all_warnings else ""
        messagebox.showinfo(
            "完了",
            f"マージしました（{len(merged)}行）。\n保存先:\n{out_path}{warn_text}",
        )
        open_in_explorer(out_path)

    def on_aol_send_click(self) -> None:
        missing: list[str] = []
        if not self.emp_picker.path:
            missing.append("社員一覧 (CSV)")
        if not self.shozoku_picker.path:
            missing.append("所属一覧 (XLSX)")
        if not self.aol_picker.path:
            missing.append("本部一覧 (XLSX)")
        if not self.tpl_picker.path:
            missing.append("個別送信テンプレート (XLSX)")
        if missing:
            messagebox.showwarning(
                "リファレンス未指定",
                "以下のリファレンスファイルを指定してください:\n・" + "\n・".join(missing),
            )
            return

        merged, all_warnings = self._merge_dropped_csvs()
        if merged is None:
            return

        try:
            emp_dir = load_employee_directory(self.emp_picker.path)
        except Exception as e:
            messagebox.showerror("読み込み失敗", f"社員一覧の読み込みに失敗しました:\n{e}")
            return
        try:
            shozoku_map = load_shozoku_map(self.shozoku_picker.path)
        except Exception as e:
            messagebox.showerror("読み込み失敗", f"所属一覧の読み込みに失敗しました:\n{e}")
            return
        try:
            aol_map = load_aol_to_cc(self.aol_picker.path)
        except Exception as e:
            messagebox.showerror("読み込み失敗", f"本部一覧の読み込みに失敗しました:\n{e}")
            return
        try:
            defaults = load_template_defaults(self.tpl_picker.path)
        except Exception as e:
            messagebox.showerror("読み込み失敗", f"個別送信テンプレートの読み込みに失敗しました:\n{e}")
            return

        # 連絡先オーバーライド（任意）: 氏名優先→本部/所属一致で本部一覧の宛先を上書き。
        contact_override = None
        if self.contact_picker.path:
            try:
                contact_override = load_contact_override(self.contact_picker.path)
            except Exception as e:
                messagebox.showerror("読み込み失敗", f"連絡先オーバーライドの読み込みに失敗しました:\n{e}")
                return

        send_df, unmatched = build_send_rows(
            merged, emp_dir, shozoku_map, aol_map, defaults, contact_override
        )

        out_dir = get_downloads_dir()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # 未マッチがあれば必ずログを残す（送信CSV出力可否に関わらず）
        skipped_path: Path | None = None
        if unmatched:
            skipped_path = out_dir / f"aol_send_skipped_{timestamp}.csv"
            try:
                write_skipped_log(unmatched, skipped_path)
            except Exception as e:
                messagebox.showwarning(
                    "ログ書き出し失敗",
                    f"未マッチログの書き出しに失敗しました（処理は継続）:\n{e}",
                )
                skipped_path = None

        if send_df.empty:
            detail = "\n".join(
                f"・{u['氏名']} [{u['種別']}] - {u['理由']}" for u in unmatched
            )
            log_line = f"\n\n未マッチログ:\n{skipped_path}" if skipped_path else ""
            messagebox.showerror(
                "出力失敗",
                f"出力可能な行がありません。\n未マッチ {len(unmatched)} 件:\n{detail}{log_line}",
            )
            if skipped_path:
                open_in_explorer(skipped_path)
            return

        out_path = out_dir / f"aol_send_{timestamp}.csv"
        try:
            send_df.to_csv(out_path, index=False, encoding="utf-8-sig")
        except Exception as e:
            messagebox.showerror("書き出し失敗", f"AOL送信CSVの書き出しに失敗しました:\n{e}")
            return

        self.status.config(text=f"出力: {out_path}")

        summary = (
            f"AOL送信CSVを出力しました（{len(send_df)}行 / マージ {len(merged)}行）。\n"
            f"保存先:\n{out_path}"
        )
        if unmatched:
            detail = "\n".join(
                f"・{u['氏名']} [{u['種別']}] - {u['理由']}" for u in unmatched
            )
            summary += f"\n\n未マッチ ({len(unmatched)}件):\n{detail}"
            if skipped_path:
                summary += f"\n\n未マッチログ:\n{skipped_path}"
        if all_warnings:
            summary += "\n\n警告:\n" + "\n".join(all_warnings)

        messagebox.showinfo("完了", summary)
        open_in_explorer(out_path)


def main() -> None:
    root = TkinterDnD.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
